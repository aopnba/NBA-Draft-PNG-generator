from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "college_draft.xlsx"
OUTPUT_PATH = ROOT / "data" / "players.json"


def to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def parse_height_to_inches(value: Any) -> float | int | None:
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric.is_integer():
            return int(numeric)
        return numeric

    if isinstance(value, str):
        match = re.match(r"^\s*(\d+)'\s*([0-9]+(?:\.[0-9]+)?)''\s*$", value)
        if match:
            feet = int(match.group(1))
            inches = float(match.group(2))
            total = (feet * 12) + inches
            if total.is_integer():
                return int(total)
            return total

    raise ValueError(f"Unsupported height value: {value!r}")


def format_height_display(value: Any) -> str | None:
    if value in (None, ""):
        return None

    if isinstance(value, str):
        match = re.match(r"^\s*(\d+)'\s*([0-9]+(?:\.[0-9]+)?)''\s*$", value)
        if match:
            feet = int(match.group(1))
            inches = match.group(2).rstrip("0").rstrip(".")
            return f"{feet}'{inches}\""
        return value.strip()

    total_inches = float(value)
    feet = int(total_inches // 12)
    inches = total_inches - (feet * 12)
    if inches.is_integer():
        inches_text = str(int(inches))
    else:
        inches_text = f"{inches:.2f}".rstrip("0").rstrip(".")
    return f"{feet}'{inches_text}\""


def format_weight_display(value: Any) -> str | None:
    if value in (None, ""):
        return None

    numeric = float(value)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.1f}".rstrip("0").rstrip(".")


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = [str(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        name = row.get("Player")
        if not name:
            continue

        rows.append(
            {
                "rank": to_int(row.get("Rk")),
                "name": str(name).strip(),
                "season": row.get("Season"),
                "team": row.get("Team"),
                "position": row.get("Pos"),
                "playerClass": row.get("Class"),
                "games": to_int(row.get("G")),
                "totals": {
                    "points": to_int(row.get("PTS")),
                    "rebounds": to_int(row.get("TRB")),
                    "assists": to_int(row.get("AST")),
                },
                "perGame": {
                    "points": to_float(row.get("PTS/Game")),
                    "rebounds": to_float(row.get("REB/Game")),
                    "assists": to_float(row.get("AST/Game")),
                },
                "heightInches": parse_height_to_inches(row.get("Height")),
                "heightDisplay": format_height_display(row.get("Height")),
                "weightLbs": to_float(row.get("Weight")),
                "weightDisplay": format_weight_display(row.get("Weight")),
            }
        )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(
            {
                "sourceWorkbook": WORKBOOK_PATH.name,
                "playerCount": len(rows),
                "players": rows,
            },
            indent=2,
        )
    )

    print(f"Wrote {len(rows)} players to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
